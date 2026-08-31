from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.application.use_cases.export_excel import ExcelExporter
from app.domain.entities import RegistroDIP

_AZUL_HEADER = "1F4E79"
_VERDE = "E2EFDA"

_HEADERS = [
    "N°", "Cama", "Servicio", "Sala", "Procedencia",
    "RUT", "Nombre Paciente", "Edad",
    "DIP", "Ubicación DIP", "Fecha Ingreso Sala", "Fecha Instalación DIP",
    "Fecha Retiro", "Días Dispositivo", "Días Hospitalización",
    "Estado", "Observaciones",
]

_COL_WIDTHS = [5, 8, 14, 22, 15, 14, 30, 7, 20, 15, 18, 20, 14, 14, 18, 12, 35]


class OpenpyxlExcelExporter(ExcelExporter):
    def export(self, records: list[RegistroDIP], titulo: str = "Vigilancia DIP") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # Excel limita a 31 chars

        encabezado = (
            f"PLANILLA VIGILANCIA DISPOSITIVOS INVASIVOS — {titulo.upper()} — "
            f"{date.today().strftime('%d/%m/%Y')}"
        )
        self._write_title(ws, encabezado)
        self._write_headers(ws)
        self._write_records(ws, records)
        self._apply_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_title(self, ws, texto: str) -> None:
        ws.merge_cells("A1:Q1")
        cell = ws["A1"]
        cell.value = texto
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

    def _write_headers(self, ws) -> None:
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_AZUL_HEADER)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 30

    def _write_records(self, ws, records: list[RegistroDIP]) -> None:
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i, record in enumerate(records):
            row = i + 3
            fill = PatternFill("solid", fgColor=_VERDE if i % 2 == 0 else "FFFFFF")
            values = [
                i + 1, record.cama, record.servicio or "", record.sala or "",
                record.procedencia or "", record.rut, record.nombre, record.edad,
                record.dip.value, record.ubicacion_dip or "N/A",
                record.fecha_ingreso_sala, record.fecha_instalacion, record.fecha_retiro,
                record.dias_dispositivo, record.dias_hospitalizacion,
                record.estado.value, record.observaciones,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

    def _apply_column_widths(self, ws) -> None:
        for col, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
